from __future__ import annotations

from openpyxl import Workbook, load_workbook
import pytest

from npl_extract.contracts import EvidenceRef, ExtractionFact, FactStatus
from npl_extract.export import export_facts


def test_exports_each_entity_to_the_42_field_template_with_evidence(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "证券代码"
    worksheet["A2"] = "资产全称"
    worksheet["A3"] = "初始起算日"
    workbook.save(template)
    evidence = EvidenceRef(
        evidence_id="p001:b001",
        artifact_scope="pypdf-all",
        document_name="发行公告.pdf",
        physical_page=1,
        locator="公告标题",
        exact_text="臻粹不良资产支持证券发行公告",
    )
    facts = [
        ExtractionFact(
            fact_id="asset-name",
            field_id="asset_full_name",
            entity_key="product:test",
            status=FactStatus.DISCLOSED,
            value="臻粹不良资产支持证券",
            evidence=[evidence],
        ),
        ExtractionFact(
            fact_id="security-code",
            field_id="security_code",
            entity_key="security:2689075",
            status=FactStatus.DISCLOSED,
            value="2689075",
            evidence=[evidence],
        ),
    ]
    output = tmp_path / "export.xlsx"

    export_facts(template, facts, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook["product_test"]["B2"].value == "臻粹不良资产支持证券"
    assert workbook["security_2689075"]["B1"].value == "2689075"
    assert workbook["security_2689075"]["B2"].value is None
    evidence_sheet = workbook["证据"]
    assert list(evidence_sheet.values)[1] == (
        "product:test",
        "asset_full_name",
        "disclosed",
        "臻粹不良资产支持证券",
        "发行公告.pdf",
        1,
        "公告标题",
        "p001:b001",
        "臻粹不良资产支持证券发行公告",
    )


def test_leaves_an_ambiguous_candidate_out_of_the_main_template(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "资产全称"
    workbook.save(template)
    fact = ExtractionFact(
        fact_id="ambiguous-name",
        field_id="asset_full_name",
        entity_key="product:test",
        status=FactStatus.AMBIGUOUS,
        value="可能的产品名称",
        evidence=[
            EvidenceRef(
                evidence_id="p001:b001",
                artifact_scope="pypdf-all",
                document_name="发行公告.pdf",
                physical_page=1,
                locator="公告标题",
                exact_text="可能的产品名称",
            )
        ],
    )
    output = tmp_path / "export.xlsx"

    export_facts(template, [fact], output)

    workbook = load_workbook(output, data_only=False)
    assert workbook["product_test"]["B1"].value is None
    assert list(workbook["证据"].values)[1][:4] == ("product:test", "asset_full_name", "ambiguous", "可能的产品名称")


def test_rejects_duplicate_exportable_values_for_one_entity_field(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "资产全称"
    workbook.save(template)
    evidence = EvidenceRef(
        evidence_id="p001:b001",
        artifact_scope="pypdf-all",
        document_name="发行公告.pdf",
        physical_page=1,
        locator="公告标题",
        exact_text="候选名称",
    )
    facts = [
        ExtractionFact(
            fact_id=f"asset-name-{index}",
            field_id="asset_full_name",
            entity_key="product:test",
            status=FactStatus.DISCLOSED,
            value=f"候选名称{index}",
            evidence=[evidence],
        )
        for index in (1, 2)
    ]

    with pytest.raises(ValueError, match="ambiguous export value"):
        export_facts(template, facts, tmp_path / "export.xlsx")


def test_exports_a_string_array_as_canonical_json_text(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "债项评级"
    workbook.save(template)
    fact = ExtractionFact(
        fact_id="ratings",
        field_id="issue_rating",
        entity_key="security:2689075",
        status=FactStatus.DISCLOSED,
        value=["中债资信:AAAsf", "中诚信国际:AAAsf"],
        evidence=[
            EvidenceRef(
                evidence_id="p002:b013",
                artifact_scope="pypdf-pages-2-2",
                document_name="发行说明书.pdf",
                physical_page=2,
                locator="发行要素/评级",
                exact_text="AAAsf/AAAsf",
            )
        ],
    )
    output = tmp_path / "export.xlsx"

    export_facts(template, [fact], output)

    workbook = load_workbook(output, data_only=False)
    assert workbook["security_2689075"]["B1"].value == '["中债资信:AAAsf","中诚信国际:AAAsf"]'
    assert list(workbook["证据"].values)[1][3] == '["中债资信:AAAsf","中诚信国际:AAAsf"]'


def test_exports_a_boolean_as_canonical_lowercase_text(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "是否含循环购买"
    workbook.save(template)
    fact = ExtractionFact(
        fact_id="static-pool",
        field_id="has_revolving_purchase",
        entity_key="product:test",
        status=FactStatus.DISCLOSED,
        value=False,
        evidence=[
            EvidenceRef(
                evidence_id="p090:b005",
                artifact_scope="pypdf-pages-90-90",
                document_name="发行说明书.pdf",
                physical_page=90,
                locator="静态池",
                exact_text="不会购买或替换资产",
            )
        ],
    )
    output = tmp_path / "export.xlsx"

    export_facts(template, [fact], output)

    workbook = load_workbook(output, data_only=False)
    assert workbook["product_test"]["B1"].value == "false"
    assert list(workbook["证据"].values)[1][3] == "false"


def test_exports_cashflow_rows_to_a_table_sheet_without_row_entity_template_tabs(tmp_path) -> None:
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    fact = ExtractionFact(
        fact_id="cashflow:2026-01",
        field_id="cashflow_collection_table",
        entity_key="cashflow_row:臻粹2026-2:2026-01",
        status=FactStatus.DISCLOSED,
        value={"period": "2026-01", "expected_recovery_amount_10k_cny": "160.70", "expected_recovery_amount_ratio_percent": "0.65"},
        evidence=[
            EvidenceRef(
                evidence_id="p112:t001:r001:c001", artifact_scope="ppstructure-v3-pages-112-113", document_name="发行说明书.pdf",
                physical_page=112, locator="资产池预计整体回收分布情况/2026-01/预计回收金额（万元）", exact_text="160.70",
            )
        ],
    )

    output = tmp_path / "export.xlsx"
    export_facts(template, [fact], output)

    workbook = load_workbook(output, data_only=False)
    assert "cashflow_row_臻粹2026-2_2026-01" not in workbook.sheetnames
    assert list(workbook["现金流归集表"].values) == [
        ("期数", "预计回收金额（万元）", "预计回收金额占比（%）", "报告名", "页码", "证据 ID"),
        ("2026-01", "160.70", "0.65", "发行说明书.pdf", 112, "p112:t001:r001:c001"),
    ]

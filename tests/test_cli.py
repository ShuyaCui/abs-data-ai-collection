from __future__ import annotations

import json
import re
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

from npl_extract.cli import main
from npl_extract.contracts import ExtractionFact, FactStatus
from npl_extract.parsers import Block, PageContent, Table, TableCell


def test_inspect_command_emits_machine_readable_result(tmp_path: Path, capsys) -> None:
    source = tmp_path / "safe.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())

    exit_code = main(["inspect", str(source)])

    result = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert result["accepted"]
    assert result["page_count"] == 1


def test_parse_command_writes_evidence_artifacts(tmp_path: Path) -> None:
    source = tmp_path / "safe.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())

    exit_code = main(["parse", str(source), "--runs-dir", str(tmp_path / "runs")])

    assert exit_code == 0
    assert list((tmp_path / "runs").glob("*/pypdf-6-16-1-all/manifest.json"))


def test_review_page_command_writes_an_offline_review_page(tmp_path: Path, capsys) -> None:
    facts = tmp_path / "candidate.jsonl"
    facts.write_text(
        json.dumps(
            {
                "fact_id": "candidate:1",
                "field_id": "asset_full_name",
                "entity_key": "product:test",
                "status": "disclosed",
                "value": "测试产品",
                "effective_at": None,
                "evidence": [],
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    fields = tmp_path / "fields.json"
    fields.write_text(json.dumps({"fields": [{"id": "asset_full_name", "export_name": "资产全称"}], "supporting_fields": []}), encoding="utf-8")
    output = tmp_path / "review.html"

    exit_code = main(["review-page", "--facts", str(facts), "--fields", str(fields), "--output", str(output)])

    assert exit_code == 0
    assert output.is_file()
    assert json.loads(capsys.readouterr().out) == {"output": str(output)}


def test_extract_command_routes_an_issuance_result_document(tmp_path: Path, capsys) -> None:
    source = tmp_path / "臻粹不良资产支持证券簿记建档发行结果公告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())

    exit_code = main(["extract", str(source), "--runs-dir", str(tmp_path / "runs")])

    assert exit_code == 3
    assert json.loads(capsys.readouterr().out) == []


def test_extract_folder_writes_candidates_and_records_unhandled_documents(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行公告.pdf"
    unhandled = tmp_path / "尚未支持的附件.pdf"
    for path in (source, unhandled):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        path.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.append(["资产全称", None])
    workbook.save(template)
    output = tmp_path.parent / "candidate.xlsx"
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [PageContent(1, "", [Block("p001:b001", 1, "臻粹不良资产支持证券发行公告", None)])],
    )

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    summary = json.loads(capsys.readouterr().out)
    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert summary["fact_count"] == 1
    assert output.is_file() and output.with_suffix(".jsonl").is_file()
    assert {item["document_name"]: item["status"] for item in manifest["documents"]} == {
        source.name: "processed", unhandled.name: "unsupported"
    }


def test_extract_folder_routes_ccxi_rating_report_pool_balance(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券信用评级报告及跟踪评级安排(中诚信国际).pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(
                4,
                "",
                [
                    Block("p004:b017", 4, "资产池特征（于初始起算日）", None),
                    Block("p004:b019", 4, "资产池未偿本息费余额 314,258.72 万元", None),
                ],
            )
        ],
    )

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    facts = [json.loads(line) for line in output.with_suffix(".jsonl").read_text().splitlines()]
    assert exit_code == 0
    assert [(fact["field_id"], fact["value"]) for fact in facts] == [
        ("initial_pool_outstanding_principal_interest_fees", "3142587200")
    ]


def test_extract_folder_persists_cashflow_table_rows_from_coordinate_cells(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"
    table = Table("p112:t001", 112, [
        TableCell("p112:t001:r000:c000", 112, "p112:t001", 0, 0, "期数", [0, 0, 1, 1]),
        TableCell("p112:t001:r000:c001", 112, "p112:t001", 0, 1, "预计回收金额（万元）", [1, 0, 2, 1]),
        TableCell("p112:t001:r000:c002", 112, "p112:t001", 0, 2, "预计回收金额占比（%）", [2, 0, 3, 1]),
        TableCell("p112:t001:r001:c000", 112, "p112:t001", 1, 0, "2026 年 1 月", [0, 1, 1, 2]),
        TableCell("p112:t001:r001:c001", 112, "p112:t001", 1, 1, "160.70", [1, 1, 2, 2]),
        TableCell("p112:t001:r001:c002", 112, "p112:t001", 1, 2, "0.65", [2, 1, 3, 2]),
        TableCell("p112:t001:r002:c000", 112, "p112:t001", 2, 0, "合计", [0, 2, 1, 3]),
        TableCell("p112:t001:r002:c001", 112, "p112:t001", 2, 1, "160.70", [1, 2, 2, 3]),
        TableCell("p112:t001:r002:c002", 112, "p112:t001", 2, 2, "0.65", [2, 2, 3, 3]),
    ])
    calls = []

    def parse(*args, **kwargs):
        calls.append((kwargs["parser"], kwargs["page_range"]))
        return [PageContent(112, "现金流归集表", tables=[table])] if kwargs["parser"] == "ppstructure" else []

    monkeypatch.setattr("npl_extract.cli.parse_native_pdf_isolated", parse)

    exit_code = main([
        "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
        "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
    ])

    facts = [json.loads(line) for line in output.with_suffix(".jsonl").read_text().splitlines()]
    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert [fact["entity_key"] for fact in facts] == ["cashflow_row:test:2026-01", "cashflow_row:test:total"]
    assert load_workbook(output)["现金流归集表"]["A2"].value == "2026-01"
    assert "cashflow_collection_table" not in manifest["field_statuses"]
    assert ("ppstructure", (112, 113)) in calls
    assert re.fullmatch(r"ppstructure-v3-paddleocr-[a-z0-9-]+-pages-112-113-[0-9a-f]{64}", facts[0]["evidence"][0]["artifact_scope"])


def test_extract_folder_records_the_field39_runtime_blocker(tmp_path: Path, capsys) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main([
        "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
        "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
    ])

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert manifest["field_statuses"]["cashflow_collection_table"] == {
        "status": "BLOCKED", "reason": "PPSTRUCTURE_NATIVE_X86_PREFLIGHT_REQUIRED"
    }


def test_extract_folder_refuses_ambiguous_duplicate_document_roles(tmp_path: Path, capsys) -> None:
    for name in ("甲产品发行公告.pdf", "乙产品发行公告.pdf"):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        (tmp_path / name).write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "甲产品", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert json.loads(capsys.readouterr().out)["fact_count"] == 0
    assert {item["status"] for item in manifest["documents"]} == {"ambiguous"}


def test_extract_folder_continues_after_one_document_parser_failure(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行公告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("PARSER_TIMEOUT: test")),
    )

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert output.is_file()
    document = manifest["documents"][0]
    assert {key: document[key] for key in ("document_name", "status", "role", "error_code")} == {
        "document_name": source.name, "status": "failed", "role": "issuance_announcement", "error_code": "PARSER_TIMEOUT"
    }
    assert document["source_sha256"] == sha256(source.read_bytes()).hexdigest()


def test_extract_folder_refuses_tied_or_unparseable_trustee_periods(tmp_path: Path, capsys) -> None:
    for name in (
        "臻粹不良资产证券化信托受托机构报告总第4期甲.pdf",
        "臻粹不良资产证券化信托受托机构报告总第4期乙.pdf",
    ):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        (tmp_path / name).write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert {item["status"] for item in manifest["documents"]} == {"ambiguous"}


def test_extract_folder_uses_latest_safe_trustee_when_newer_file_is_rejected(tmp_path: Path, capsys) -> None:
    safe = tmp_path / "臻粹不良资产证券化信托受托机构报告总第3期.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    safe.write_bytes(content.getvalue())
    rejected = tmp_path / "臻粹不良资产证券化信托受托机构报告总第4期.pdf"
    rejected.write_bytes(b"%PDF-not-a-valid-pdf")
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert {item["document_name"]: item["status"] for item in manifest["documents"]} == {
        safe.name: "no_facts", rejected.name: "rejected"
    }


def test_extract_folder_refuses_mixed_product_document_names(tmp_path: Path, capsys) -> None:
    for name in ("甲产品不良资产发行公告.pdf", "乙产品不良资产发行说明书.pdf"):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        (tmp_path / name).write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "甲产品不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert {item["status"] for item in manifest["documents"]} == {"ambiguous"}


def test_extract_folder_refuses_mixed_products_hidden_by_duplicate_roles(tmp_path: Path, capsys) -> None:
    for name in ("甲产品不良资产发行公告.pdf", "乙产品不良资产发行公告.pdf", "乙产品不良资产发行说明书.pdf"):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        (tmp_path / name).write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "甲产品不良资产", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert {item["status"] for item in manifest["documents"]} == {"ambiguous"}


def test_extract_folder_distinguishes_product_terms_after_npl_label(tmp_path: Path, capsys) -> None:
    for name in ("甲不良资产支持证券第一期发行公告.pdf", "甲不良资产支持证券第二期发行说明书.pdf"):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        (tmp_path / name).write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "甲不良资产第一期", "--template", str(template),
            "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
        ]
    )

    manifest = json.loads(output.with_suffix(".manifest.json").read_text())
    assert exit_code == 0
    assert {item["status"] for item in manifest["documents"]} == {"ambiguous"}


def test_extract_folder_rejects_template_as_output(tmp_path: Path, capsys) -> None:
    template = tmp_path.parent / "template.xlsx"
    Workbook().save(template)

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "test", "--template", str(template),
            "--output", str(template),
        ]
    )

    assert exit_code == 2
    assert "outside the input directory" in json.loads(capsys.readouterr().out)["error"]


def test_extract_folder_rejects_an_output_held_by_another_job(tmp_path: Path, capsys, monkeypatch) -> None:
    template = tmp_path.parent / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"
    monkeypatch.setattr("npl_extract.cli._exclusive_output_lock", lambda path: (_ for _ in ()).throw(BlockingIOError()))

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "test", "--template", str(template),
            "--output", str(output),
        ]
    )

    assert exit_code == 2
    assert "already in use" in json.loads(capsys.readouterr().out)["error"]


def test_extract_folder_rejects_output_inside_the_input_folder(tmp_path: Path, capsys) -> None:
    template = tmp_path / "template.xlsx"
    Workbook().save(template)

    exit_code = main(
        [
            "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "test", "--template", str(template),
            "--output", str(tmp_path / "candidate.xlsx"),
        ]
    )

    assert exit_code == 2
    assert "outside the input directory" in json.loads(capsys.readouterr().out)["error"]


def test_extract_command_persists_an_issuance_announcement_fact_set(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行公告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(1, "", [Block("p001:b001", 1, "臻粹不良资产支持证券发行公告", [0, 0, 72, 72])])
        ],
    )

    exit_code = main(["extract", str(source), "--entity-key", "product:test", "--runs-dir", str(tmp_path / "runs")])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output[0]["field_id"] == "asset_full_name"
    manifest = next((tmp_path / "runs").glob("*/pypdf-*-all/manifest.json"))
    facts_path = next((tmp_path / "runs").glob("*/facts/*.jsonl"))
    assert json.loads(facts_path.read_text())["evidence"][0]["artifact_scope"] == json.loads(manifest.read_text())["scope"]


def test_extract_command_persists_an_ocr_issuance_result_fact_set(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券簿记建档发行结果公告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(
                1,
                "",
                [
                    Block("p001:b001", 1, "臻粹不良资产支持证券簿记建档发行结果公告", [0, 0, 72, 72]),
                    Block("p001:b009", 1, "证券代码", [0, 0, 72, 72]),
                    Block("p001:b010", 1, "2689075", [0, 0, 72, 72]),
                    Block("p001:b011", 1, "预期到期日", [0, 0, 72, 72]),
                    Block("p001:b012", 1, "2028年2月23日", [0, 0, 72, 72]),
                    Block("p001:b019", 1, "实际发行总额", [0, 0, 72, 72]),
                    Block("p001:b020", 1, "13,200.00万元", [0, 0, 72, 72]),
                ],
                ocr_requested=True,
            )
        ],
    )

    exit_code = main(["extract", str(source), "--native-parser", "docling-ocr", "--runs-dir", str(tmp_path / "runs")])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert {fact["field_id"] for fact in output} == {"security_code", "maturity_date", "tranche_issue_amount"}
    assert next((tmp_path / "runs").glob("*/docling-*-all/manifest.json")).is_file()
    assert next((tmp_path / "runs").glob("*/facts/*.jsonl")).is_file()


def test_extract_command_persists_prospectus_issue_amounts_without_associations(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(
                2,
                "",
                [
                    Block("p002:b001", 2, "证券名称 发行金额（万元）规模占比", None),
                    Block("p002:b002", 2, "优先档 13,200.00 72.53% 过手", None),
                    Block("p002:b003", 2, "次级档 5,000.00 27.47% 过手", None),
                    Block("p002:b004", 2, "总计 18,200.00 100.00% -", None),
                ],
            )
        ],
    )

    exit_code = main(["extract", str(source), "--entity-key", "product:test", "--runs-dir", str(tmp_path / "runs")])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert {(fact["field_id"], fact["value"]) for fact in output} == {
        ("issue_amount_senior", "1.32"),
        ("issue_amount_mezzanine", None),
        ("issue_amount_subordinated", "0.5"),
    }


def test_extract_command_persists_prospectus_market_terms_without_associations(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(3, "", [Block("p003:b001", 3, "本期资产支持证券拟采用公开簿记建档的方式在全国银行间债券市场发行。", None)])
        ],
    )

    exit_code = main(["extract", str(source), "--entity-key", "product:test", "--runs-dir", str(tmp_path / "runs")])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert {(fact["field_id"], fact["value"]) for fact in output} == {
        ("market", "银行间债券市场"),
        ("issuance_method", "簿记建档"),
    }


def test_extract_command_persists_a_participant_list_financing_entity(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(
                16,
                "",
                [
                    Block("p016:b020", 16, "二、各参与机构名单", None),
                    Block("p016:b021", 16, "发起机构/贷款服务机构：广发银行股份有限公司（简称广发银行）", None),
                ],
            )
        ],
    )

    exit_code = main(["extract", str(source), "--entity-key", "product:test", "--runs-dir", str(tmp_path / "runs")])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert [(fact["field_id"], fact["value"]) for fact in output] == [("actual_financing_entity", ["广发银行股份有限公司"])]


def test_extract_command_rejects_a_security_key_for_product_facts(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹不良资产支持证券发行公告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(1, "", [Block("p001:b001", 1, "臻粹不良资产支持证券发行公告", [0, 0, 72, 72])])
        ],
    )

    exit_code = main(["extract", str(source), "--entity-key", "security:2689075", "--runs-dir", str(tmp_path / "runs")])

    assert exit_code == 3
    assert json.loads(capsys.readouterr().out) == []


def test_extract_command_projects_prospectus_payment_date_through_association_facts(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "臻粹2026年第二期不良资产支持证券发行说明书.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    associations = tmp_path / "associations.jsonl"
    associations.write_text(
        json.dumps(
            {
                "fact_id": "senior-level",
                "field_id": "tranche_level",
                "entity_key": "security:2689075",
                "status": "disclosed",
                "value": "优先档",
                "evidence": [
                    {
                        "evidence_id": "p001:b005",
                        "artifact_scope": "docling-ocr-all",
                        "document_name": "臻粹2026年第二期不良资产支持证券簿记建档发行结果公告.pdf",
                        "physical_page": 1,
                        "locator": "证券名称",
                        "exact_text": "优先档",
                    }
                ],
            },
            ensure_ascii=False,
        )
        + "\n"
    )
    monkeypatch.setattr(
        "npl_extract.cli.parse_native_pdf_isolated",
        lambda *args, **kwargs: [
            PageContent(
                2,
                "",
                [
                    Block("p002:b028", 2, "资产支持证券的第一个支付日是 2026 年 5 月 23 日", None),
                    Block("p002:b011", 2, "方式 利率类型 预期到期日 法定到期日 评级", None),
                    Block("p002:b012", 2, "（中债资信/中诚信）", None),
                    Block("p002:b013", 2, "优先档 13,200.00 72.53% 过手 固定利率 2028/2/23 2032/4/23 AAAsf/AAAsf", None),
                ],
            )
        ],
    )

    exit_code = main(
        [
            "extract",
            str(source),
            "--entity-key",
            "product:test",
            "--association-facts",
            str(associations),
            "--runs-dir",
            str(tmp_path / "runs"),
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert {fact["field_id"] for fact in output} == {"first_interest_payment_date", "issue_rating"}
    assert {fact["entity_key"] for fact in output} == {"security:2689075"}


def test_extract_trustee_command_rejects_a_product_key_for_report_facts(tmp_path: Path, capsys, monkeypatch) -> None:
    source = tmp_path / "受托报告.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    content = BytesIO()
    writer.write(content)
    source.write_bytes(content.getvalue())
    monkeypatch.setattr(
        "npl_extract.cli.extract_trustee_report_facts",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("extractor must not run")),
    )

    exit_code = main(["extract-trustee", str(source), "--entity-key", "product:test", "--runs-dir", str(tmp_path / "runs")])

    assert exit_code == 3
    assert json.loads(capsys.readouterr().out) == []


def test_export_command_projects_persisted_facts_to_a_template(tmp_path: Path, capsys) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "资产全称"
    workbook.save(template)
    facts = tmp_path / "facts.jsonl"
    facts.write_text(
        json.dumps(
            {
                "fact_id": "asset-name",
                "field_id": "asset_full_name",
                "entity_key": "product:test",
                "status": "disclosed",
                "value": "臻粹不良资产支持证券",
                "evidence": [
                    {
                        "evidence_id": "p001:b001",
                        "artifact_scope": "pypdf-all",
                        "document_name": "发行公告.pdf",
                        "physical_page": 1,
                        "locator": "公告标题",
                        "exact_text": "臻粹不良资产支持证券发行公告",
                    }
                ],
                "effective_at": None,
                "rule_version": None,
                "derived_inputs": [],
                "confirmed": False,
            },
            ensure_ascii=False,
        )
        + "\n"
    )
    output = tmp_path / "export.xlsx"

    exit_code = main(["export", "--template", str(template), "--facts", str(facts), "--output", str(output)])

    assert exit_code == 0
    assert json.loads(capsys.readouterr().out)["fact_count"] == 1
    assert load_workbook(output)["product_test"]["B1"].value == "臻粹不良资产支持证券"


def test_export_command_rejects_a_jsonl_fact_with_the_wrong_entity_grain(tmp_path: Path, capsys) -> None:
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    facts = tmp_path / "facts.jsonl"
    facts.write_text(
        json.dumps(
            {
                "fact_id": "asset-name",
                "field_id": "asset_full_name",
                "entity_key": "security:2689075",
                "status": "disclosed",
                "value": "臻粹不良资产支持证券",
                "evidence": [
                    {
                        "evidence_id": "p001:b001",
                        "artifact_scope": "pypdf-all",
                        "document_name": "发行公告.pdf",
                        "physical_page": 1,
                        "locator": "公告标题",
                        "exact_text": "臻粹不良资产支持证券发行公告",
                    }
                ],
            },
            ensure_ascii=False,
        )
        + "\n"
    )

    exit_code = main(["export", "--template", str(template), "--facts", str(facts), "--output", str(tmp_path / "export.xlsx")])

    assert exit_code == 2
    assert "product key" in json.loads(capsys.readouterr().out)["error"]


def test_review_command_appends_an_accepted_fact_decision(tmp_path: Path, capsys) -> None:
    document_sha256 = "a" * 64
    content = (
        json.dumps(
            {
                "fact_id": "asset-name",
                "field_id": "asset_full_name",
                "entity_key": "product:test",
                "status": "disclosed",
                "value": "臻粹不良资产支持证券",
                "evidence": [
                    {
                        "evidence_id": "p001:b001",
                        "artifact_scope": "pypdf-all",
                        "document_name": "发行公告.pdf",
                        "physical_page": 1,
                        "locator": "公告标题",
                        "exact_text": "臻粹不良资产支持证券发行公告",
                    }
                ],
            },
            ensure_ascii=False,
        )
        + "\n"
    )
    runs_dir = tmp_path / "runs"
    candidates = runs_dir / document_sha256 / "facts" / f"{sha256(content.encode()).hexdigest()}.jsonl"
    candidates.parent.mkdir(parents=True)
    candidates.write_text(content)

    exit_code = main(
        [
            "review",
            "--document-sha256",
            document_sha256,
            "--facts",
            str(candidates),
            "--fact-id",
            "asset-name",
            "--action",
            "accept",
            "--decision-id",
            "review-001",
            "--reviewer-id",
            "business-owner:alice",
            "--reason-code",
            "VALUE_AND_EVIDENCE_CONFIRMED",
            "--runs-dir",
            str(runs_dir),
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["action"] == "accept"
    assert output["resolved_fact"]["confirmed"] is True
    assert (runs_dir / document_sha256 / "reviews" / "review-001.json").is_file()


def test_review_command_accepts_a_derived_fact_after_its_inputs_are_reviewed(tmp_path: Path, capsys) -> None:
    document_sha256 = "c" * 64
    runs_dir = tmp_path / "runs"
    evidence = {"evidence_id": "p007:b001", "artifact_scope": "pypdf-pages-1-7", "document_name": "受托机构报告.pdf", "physical_page": 7, "locator": "回收", "exact_text": "30,466,642.99"}
    facts = [
        {"fact_id": "in-progress", "field_id": "npl_recovery_in_progress_cumulative", "entity_key": "report:test", "status": "disclosed", "value": "30466642.99", "evidence": [evidence]},
        {"fact_id": "completed", "field_id": "npl_recovery_completed_cumulative", "entity_key": "report:test", "status": "disclosed", "value": "29941313.75", "evidence": [evidence]},
        {"fact_id": "derived-recovery", "field_id": "npl_trustee_recovery_cash", "entity_key": "report:test", "status": "derived", "value": "0.6040795674", "evidence": [evidence], "rule_version": "npl-recovery-cash-v1", "derived_inputs": [{"fact_id": "in-progress", "confirmed": False}, {"fact_id": "completed", "confirmed": False}]},
    ]
    content = "".join(json.dumps(fact) + "\n" for fact in facts)
    candidates = runs_dir / document_sha256 / "facts" / f"{sha256(content.encode()).hexdigest()}.jsonl"
    candidates.parent.mkdir(parents=True)
    candidates.write_text(content)

    for fact_id, decision_id in (("in-progress", "review-input-1"), ("completed", "review-input-2"), ("derived-recovery", "review-derived")):
        exit_code = main([
            "review", "--document-sha256", document_sha256, "--facts", str(candidates), "--fact-id", fact_id,
            "--action", "accept", "--decision-id", decision_id, "--reviewer-id", "business-owner:alice",
            "--reason-code", "VALUE_AND_EVIDENCE_CONFIRMED", "--runs-dir", str(runs_dir),
        ])
        assert exit_code == 0, capsys.readouterr().out
        capsys.readouterr()

    decision = json.loads((runs_dir / document_sha256 / "reviews" / "review-derived.json").read_text())
    assert all(item["confirmed"] for item in decision["resolved_fact"]["derived_inputs"])


def test_review_command_rejects_a_candidate_from_another_document(tmp_path: Path, capsys) -> None:
    runs_dir = tmp_path / "runs"
    candidates = runs_dir / ("b" * 64) / "facts" / f"{sha256(b'').hexdigest()}.jsonl"
    candidates.parent.mkdir(parents=True)
    candidates.write_text("")

    exit_code = main(
        [
            "review",
            "--document-sha256",
            "a" * 64,
            "--facts",
            str(candidates),
            "--fact-id",
            "asset-name",
            "--action",
            "accept",
            "--decision-id",
            "review-002",
            "--reviewer-id",
            "business-owner:alice",
            "--reason-code",
            "VALUE_AND_EVIDENCE_CONFIRMED",
            "--runs-dir",
            str(runs_dir),
        ]
    )

    assert exit_code == 2
    assert "canonical" in json.loads(capsys.readouterr().out)["error"]


def test_review_command_rejects_a_tampered_candidate_artifact(tmp_path: Path, capsys) -> None:
    runs_dir = tmp_path / "runs"
    candidates = runs_dir / ("a" * 64) / "facts" / f"{'b' * 64}.jsonl"
    candidates.parent.mkdir(parents=True)
    candidates.write_text("")

    exit_code = main(
        [
            "review",
            "--document-sha256",
            "a" * 64,
            "--facts",
            str(candidates),
            "--fact-id",
            "asset-name",
            "--action",
            "accept",
            "--decision-id",
            "review-003",
            "--reviewer-id",
            "business-owner:alice",
            "--reason-code",
            "VALUE_AND_EVIDENCE_CONFIRMED",
            "--runs-dir",
            str(runs_dir),
        ]
    )

    assert exit_code == 2
    assert "content-addressed" in json.loads(capsys.readouterr().out)["error"]


def test_extract_folder_routes_recovery_pages_after_rating_and_preserves_non_conflicting_prospectus_facts(
    tmp_path: Path, capsys, monkeypatch
) -> None:
    prospectus = tmp_path / "臻粹不良资产支持证券发行说明书.pdf"
    rating_report = tmp_path / "臻粹不良资产支持证券信用评级报告及跟踪评级安排(中诚信国际).pdf"
    for path in (prospectus, rating_report):
        writer = PdfWriter()
        writer.add_blank_page(width=72, height=72)
        content = BytesIO()
        writer.write(content)
        path.write_bytes(content.getvalue())
    template = tmp_path / "template.xlsx"
    Workbook().save(template)
    output = tmp_path.parent / "candidate.xlsx"
    calls = []

    def fact(field_id: str, value: str | list[str], source: str) -> ExtractionFact:
        return ExtractionFact(
            fact_id=f"{source}:{field_id}", field_id=field_id, entity_key="product:test", status=FactStatus.DISCLOSED,
            value=value, evidence=[{
                "evidence_id": f"{source}:b001", "artifact_scope": "pypdf-test", "document_name": source,
                "physical_page": 1, "locator": "test", "exact_text": "test",
            }],
        )

    def parse(*args, **kwargs):
        calls.append((Path(args[0]).name, kwargs["parser"], kwargs["page_range"]))
        return [PageContent(kwargs["page_range"][0], "", [])]

    monkeypatch.setattr("npl_extract.cli.parse_native_pdf_isolated", parse)
    monkeypatch.setattr(
        "npl_extract.cli.extract_rating_report_facts",
        lambda *args: [fact("chinabond_predicted_recovery_rate", "8.00", "rating")],
    )
    monkeypatch.setattr(
        "npl_extract.cli.extract_prospectus_recovery_prediction_facts",
        lambda *args: [
            fact("chinabond_predicted_recovery_rate", "7.90", "prospectus"),
            fact("chinabond_predicted_recovery_amount", "2.482766", "prospectus"),
            fact("issuance_cashflow_forecast_agency", ["中债资信", "中诚信国际"], "prospectus"),
        ],
    )

    exit_code = main([
        "extract-folder", str(tmp_path), "--product-key", "product:test", "--product-name", "臻粹不良资产", "--template", str(template),
        "--output", str(output), "--runs-dir", str(tmp_path / "runs"),
    ])

    facts = [json.loads(line) for line in output.with_suffix(".jsonl").read_text().splitlines()]
    assert exit_code == 0, capsys.readouterr().out
    assert [(parser, page_range) for _, parser, page_range in calls if page_range in {(4, 4), (102, 104)}] == [
        ("pypdf", (4, 4)),
        ("pypdf", (102, 104)),
    ]
    assert {item["field_id"]: item["value"] for item in facts} == {
        "chinabond_predicted_recovery_rate": "8.00",
        "chinabond_predicted_recovery_amount": "2.482766",
        "issuance_cashflow_forecast_agency": ["中债资信", "中诚信国际"],
    }

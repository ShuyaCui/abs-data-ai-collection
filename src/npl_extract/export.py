from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from npl_extract.contracts import ExtractionFact, FactStatus, load_field_contracts


def export_facts(template: Path, facts: list[ExtractionFact], output: Path) -> None:
    """Project unambiguous candidate facts to a 42-field template and evidence sheet."""
    workbook = load_workbook(template)
    template_sheet = workbook.active
    grouped = defaultdict(list)
    for fact in facts:
        grouped[fact.entity_key].append(fact)
    contracts = load_field_contracts()
    used_names = set()
    sheets = []
    for index, (entity_key, entity_facts) in enumerate(sorted(grouped.items())):
        sheet = template_sheet if index == 0 else workbook.copy_worksheet(template_sheet)
        sheet.title = _sheet_name(entity_key, used_names)
        sheets.append((sheet, entity_facts))
    for sheet, entity_facts in sheets:
        rows = {sheet.cell(row, 1).value: row for row in range(1, sheet.max_row + 1)}
        for fact in entity_facts:
            contract = contracts.get(fact.field_id)
            row = rows.get(contract.export_name) if contract else None
            if row is None or fact.value is None or fact.status not in {FactStatus.DISCLOSED, FactStatus.DERIVED}:
                continue
            if sum(item.field_id == fact.field_id for item in entity_facts) != 1:
                raise ValueError(f"ambiguous export value for {entity_key}/{fact.field_id}")
            cell = sheet.cell(row, 2, fact.value)
            cell.number_format = "@"
            cell.font = Font(color="0000FF")
    if not grouped:
        template_sheet.title = "未分配"
    if "证据" in workbook.sheetnames:
        del workbook["证据"]
    evidence_sheet = workbook.create_sheet("证据")
    evidence_sheet.append(["实体", "字段", "状态", "值", "报告名", "页码", "表/段落", "证据 ID", "原文"])
    for fact in facts:
        for evidence in fact.evidence:
            evidence_sheet.append(
                [
                    fact.entity_key,
                    fact.field_id,
                    fact.status.value,
                    fact.value,
                    evidence.document_name,
                    evidence.physical_page,
                    evidence.locator,
                    evidence.evidence_id,
                    evidence.exact_text,
                ]
            )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def _sheet_name(entity_key: str, used_names: set[str]) -> str:
    base = entity_key.replace(":", "_")[:31] or "未分配"
    name = base
    suffix = 1
    while name in used_names:
        suffix += 1
        name = f"{base[: 31 - len(str(suffix)) - 1]}_{suffix}"
    used_names.add(name)
    return name
